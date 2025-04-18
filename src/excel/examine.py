import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def print_sheet_summary(sheet_name, df):
    print(f"\n{'='*50}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*50}")
    print(f"Shape: {df.shape} (rows, columns)")
    
    # Identify columns with actual data (non-empty)
    non_empty_cols = []
    for col in df.columns:
        if df[col].notna().any():
            non_empty_cols.append(col)
    
    print(f"Non-empty columns: {len(non_empty_cols)} out of {len(df.columns)}")
    
    # Show first meaningful rows (first rows that have substantial data)
    meaningful_rows = df.loc[df.notna().sum(axis=1) > 2].head(10)
    if not meaningful_rows.empty:
        print("\nSample meaningful data:")
        print(meaningful_rows)
    
    # Print column headers with data
    print("\nColumn headers with data:")
    for col in non_empty_cols:
        print(f"  - {col}")

# Get basic information from pandas
print("ANALYZING EXCEL FILE STRUCTURE...")
df_dict = pd.read_excel('DF format.xlsx', sheet_name=None)
for sheet_name, df in df_dict.items():
    print_sheet_summary(sheet_name, df)

# Check for formulas using openpyxl
print("\n\nCHECKING FOR FORMULAS...")
wb = openpyxl.load_workbook('DF format.xlsx', data_only=False)

for sheet_name in wb.sheetnames:
    print(f"\n{'='*50}")
    print(f"FORMULAS IN SHEET: {sheet_name}")
    print(f"{'='*50}")
    
    ws = wb[sheet_name]
    formula_count = 0
    
    # Sample some formulas (up to 10)
    formula_examples = []
    
    for row in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
        for col in range(1, min(ws.max_column + 1, 50)):  # Check first 50 columns
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
                if len(formula_examples) < 10:
                    col_letter = get_column_letter(col)
                    formula_examples.append(f"Cell {col_letter}{row}: {cell.value}")
    
    print(f"Total formulas found: {formula_count}")
    if formula_examples:
        print("Formula examples:")
        for example in formula_examples:
            print(f"  {example}")
            
print("\nANALYSIS COMPLETE") 